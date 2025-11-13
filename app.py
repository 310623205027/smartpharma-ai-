from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime, timedelta
from dotenv import load_dotenv
import os
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Import custom modules
from database import Database
from chatbot import PharmacyChatbot
from pyzbar.pyzbar import decode
from PIL import Image

load_dotenv()

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__, template_folder='templates', static_folder='static')
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['JSON_SORT_KEYS'] = False

# Initialize database and chatbot
db = Database()
chatbot = PharmacyChatbot(db)

try:
    db.connect()
    db.create_tables()
    db.add_sample_data()
    logger.info("Database initialized successfully")
except Exception as e:
    logger.error(f"Failed to initialize database: {e}")

# ============================================================================
# MAIN ROUTES
# ============================================================================

@app.route('/')
def index():
    return render_template('dashboard.html')

@app.route('/upload')
def upload_page():
    return render_template('upload.html')

@app.route('/alerts')
def alerts_page():
    return render_template('alerts.html')

@app.route('/insights')
def insights_page():
    return render_template('insights.html')

@app.route('/sales_counter')
def sales_counter():
    return render_template('sales_counter.html')

# ============================================================================
# HEALTH CHECK
# ============================================================================

@app.route('/api/health')
def health_check():
    try:
        products = db.get_all_products()
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'products_count': len(products) if products else 0
        })
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return jsonify({'status': 'unhealthy', 'error': str(e)}), 500

# ============================================================================
# DASHBOARD API
# ============================================================================

@app.route('/api/dashboard')
def get_dashboard_data():
    try:
        products = db.get_all_products()
        
        if not products:
            return jsonify({
                'status': 'success',
                'metrics': {
                    'total_products': 0,
                    'total_stock': 0,
                    'avg_eco_score': 0,
                    'expiring': 0
                },
                'products': []
            })
        
        total_products = len(products)
        total_stock = sum(int(p.get('stock_quantity', 0)) for p in products)
        avg_eco_score = sum(float(p.get('eco_score', 5)) for p in products) / total_products
        
        expiring = db.get_expiring_products(days=7)
        expiring_count = len(expiring) if expiring else 0
        
        logger.info(f"Dashboard: {total_products} products, {total_stock} total stock")
        
        return jsonify({
            'status': 'success',
            'metrics': {
                'total_products': total_products,
                'total_stock': total_stock,
                'avg_eco_score': round(avg_eco_score, 2),
                'expiring': expiring_count
            },
            'products': products[:8]
        })
    except Exception as e:
        logger.error(f"Dashboard error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# PRODUCTS API
# ============================================================================

@app.route('/api/products')
def get_products():
    try:
        products = db.get_all_products()
        return jsonify({
            'status': 'success',
            'count': len(products) if products else 0,
            'products': products or []
        })
    except Exception as e:
        logger.error(f"Get products error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/add-product', methods=['POST'])
def add_product():
    try:
        data = request.json
        required = ['name', 'category', 'barcode', 'expiry_date', 'packaging_type', 'stock_quantity', 'price']
        
        if not all(k in data for k in required):
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400
        
        product_id = db.insert_product(
            name=data['name'],
            category=data['category'],
            barcode=data['barcode'],
            expiry_date=data['expiry_date'],
            packaging_type=data['packaging_type'],
            eco_score=data.get('eco_score', 5.0),
            stock_quantity=data['stock_quantity'],
            price=data['price'],
            mfg_date=data.get('mfg_date')
        )
        
        if product_id:
            return jsonify({
                'status': 'success',
                'product_id': product_id,
                'message': 'Product added successfully'
            })
        else:
            return jsonify({'status': 'error', 'message': 'Barcode already exists'}), 400
    except Exception as e:
        logger.error(f"Add product error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# BARCODE SCAN & UPLOAD
# ============================================================================

@app.route('/api/upload', methods=['POST'])
def upload_barcode():
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        img = Image.open(file.stream)
        decoded_objects = decode(img)
        
        if not decoded_objects:
            return jsonify({'status': 'error', 'message': 'No barcode found in image'}), 400
        
        barcode_data = decoded_objects[0].data.decode('utf-8')
        
        # Try to fetch product by barcode
        product = db.get_product_by_barcode(barcode_data)
        
        return jsonify({
            'status': 'success',
            'barcode': barcode_data,
            'product': product,
            'message': 'Barcode scanned successfully'
        })
    except Exception as e:
        logger.error(f"Upload error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# ALERTS API
# ============================================================================

@app.route('/api/alerts')
def get_alerts():
    try:
        alerts = []
        alert_id = 1
        
        # Get expiring products
        expiring_products = db.get_expiring_products(days=7)
        if expiring_products:
            for product in expiring_products:
                try:
                    expiry_date = datetime.strptime(str(product.get('expiry_date', '')), '%Y-%m-%d').date()
                    days_left = (expiry_date - datetime.now().date()).days
                    severity = 'critical' if days_left < 2 else 'warning'
                    
                    alerts.append({
                        'alert_id': alert_id,
                        'type': 'expiry',
                        'severity': severity,
                        'product': product['name'],
                        'product_id': product['id'],
                        'message': f"Expires in {days_left} days",
                        'details': product
                    })
                    alert_id += 1
                except Exception as e:
                    logger.error(f"Error processing expiry alert: {e}")
        
        # Get low stock products
        low_stock_products = db.get_low_stock_products(threshold=50)
        if low_stock_products:
            for product in low_stock_products:
                alerts.append({
                    'alert_id': alert_id,
                    'type': 'stock',
                    'severity': 'warning',
                    'product': product['name'],
                    'product_id': product['id'],
                    'message': f"Low stock: {product['stock_quantity']} units",
                    'details': product
                })
                alert_id += 1
        
        logger.info(f"Alerts fetched: {len(alerts)} total")
        
        return jsonify({
            'status': 'success',
            'alerts': alerts,
            'count': len(alerts),
            'critical': sum(1 for a in alerts if a['severity'] == 'critical'),
            'warning': sum(1 for a in alerts if a['severity'] == 'warning')
        })
    except Exception as e:
        logger.error(f"Alerts error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# INSIGHTS API
# ============================================================================

@app.route('/api/insights')
def get_insights():
    try:
        products = db.get_all_products()
        expiring = db.get_expiring_products(days=7)
        low_stock = db.get_low_stock_products(threshold=50)
        
        if not products:
            return jsonify({
                'status': 'success',
                'insights': {
                    'total_products': 0,
                    'expiring_soon': 0,
                    'low_stock': 0,
                    'avg_eco_score': 0,
                    'waste_prevented': 0,
                    'top_products': [],
                    'recommendations': []
                }
            })
        
        total_products = len(products)
        avg_eco = sum(float(p.get('eco_score', 5)) for p in products) / total_products
        total_inventory_value = sum(float(p.get('price', 0)) * int(p.get('stock_quantity', 0)) for p in products)
        
        # Top products by stock
        top_products = sorted(products, key=lambda x: int(x.get('stock_quantity', 0)), reverse=True)[:5]
        top_products_list = [{'name': p['name'], 'stock': p['stock_quantity']} for p in top_products]
        
        logger.info("Insights generated successfully")
        
        return jsonify({
            'status': 'success',
            'insights': {
                'total_products': total_products,
                'expiring_soon': len(expiring) if expiring else 0,
                'low_stock': len(low_stock) if low_stock else 0,
                'avg_eco_score': round(avg_eco, 2),
                'total_inventory_value': round(total_inventory_value, 2),
                'waste_prevented': round(len(products) * 0.15, 2),
                'top_products': top_products_list,
                'recommendations': [
                    'Review expiring stock for clearance sales',
                    'Reorder low stock items to maintain availability',
                    'Promote products with high eco-scores',
                    'Analyze seasonal demand patterns',
                    'Consider supplier negotiations for bulk orders'
                ]
            }
        })
    except Exception as e:
        logger.error(f"Insights error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# SALES COUNTER API
# ============================================================================

@app.route('/api/product/<barcode>')
def get_product(barcode):
    try:
        product = db.get_product_by_barcode(barcode)
        
        if product:
            return jsonify({
                'success': True,
                'product': product
            })
        
        return jsonify({'success': False, 'message': 'Product not found'}), 404
    except Exception as e:
        logger.error(f"Get product error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/record_sale', methods=['POST'])
def record_sale():
    try:
        data = request.json
        product_id = data.get('product_id')
        quantity = int(data.get('quantity', 0))
        
        if quantity <= 0:
            return jsonify({'success': False, 'message': 'Invalid quantity'}), 400
        
        product = db.get_product_by_id(product_id)
        
        if not product or product['stock_quantity'] < quantity:
            return jsonify({'success': False, 'message': 'Insufficient stock'}), 400
        
        # Update stock
        new_stock = db.update_stock(product_id, -quantity)
        
        if new_stock is not None:
            total_price = float(product['price']) * quantity
            return jsonify({
                'success': True,
                'total_price': total_price,
                'new_stock': new_stock
            })
        
        return jsonify({'success': False, 'message': 'Failed to record sale'}), 500
    except Exception as e:
        logger.error(f"Record sale error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/sales_report')
def sales_report():
    try:
        products = db.get_all_products()
        
        if not products:
            return jsonify({'status': 'error', 'message': 'No sales data available'}), 400
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        headers = ['Product Name', 'Barcode', 'Category', 'Stock', 'Price', 'Total Value', 'Expiry Date']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='3A9BDC', end_color='3A9BDC', fill_type='solid')
        
        for row, product in enumerate(products, 2):
            total_value = float(product.get('price', 0)) * int(product.get('stock_quantity', 0))
            ws.cell(row, 1, product['name'])
            ws.cell(row, 2, product.get('barcode', 'N/A'))
            ws.cell(row, 3, product.get('category', 'N/A'))
            ws.cell(row, 4, product['stock_quantity'])
            ws.cell(row, 5, f"₹{float(product.get('price', 0)):.2f}")
            ws.cell(row, 6, f"₹{total_value:.2f}")
            ws.cell(row, 7, product.get('expiry_date', 'N/A'))
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Sales report generated")
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'inventory_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Sales report error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# ============================================================================
# CHATBOT API (INTEGRATED)
# ============================================================================

@app.route('/api/chat', methods=['POST'])
def chat():
    try:
        data = request.json
        user_message = data.get('message', '').strip()
        
        if not user_message:
            return jsonify({
                'status': 'error',
                'response': 'Please ask me something'
            }), 400
        
        # Get response from chatbot
        response = chatbot.get_response(user_message)
        
        logger.info(f"Chat - User: {user_message[:50]}... Bot: {response[:50]}...")
        
        return jsonify({
            'status': 'success',
            'response': response
        })
    except Exception as e:
        logger.error(f"Chat error: {e}")
        return jsonify({
            'status': 'error',
            'response': f'Error: {str(e)}'
        }), 500

@app.route('/api/chat/quick-responses')
def quick_responses():
    return jsonify({
        'status': 'success',
        'responses': [
            'How many products?',
            'What\'s expiring?',
            'Low stock items?',
            'Inventory status?',
            'Category info?',
            'Eco-score analysis?'
        ]
    })

# ============================================================================
# ERROR HANDLERS
# ============================================================================

@app.errorhandler(404)
def not_found(e):
    return jsonify({'status': 'error', 'message': 'Endpoint not found'}), 404

@app.errorhandler(500)
def server_error(e):
    logger.error(f"Server error: {e}")
    return jsonify({'status': 'error', 'message': 'Internal server error'}), 500
@app.route('/chat')
def chat_page():
    return render_template('chatbot.html')
# ============================================================================
# SHUTDOWN HANDLER
# ============================================================================

@app.teardown_appcontext
def close_db(error):
    try:
        db.close()
    except Exception as e:
        logger.error(f"Error closing database: {e}")

if __name__ == '__main__':
    logger.info("Starting SmartPharma Application")
    app.run(debug=True, host='0.0.0.0', port=5000)